from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import csv
import re
import uuid
import math
import os
import json
import sqlite3
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None  # Fallback if not available
import shutil

from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
from openpyxl import Workbook, load_workbook


app = Flask(__name__)
app.secret_key = "order-stock-secret-key"

# Database configuration
DATABASE_PATH = os.path.join(os.path.dirname(__file__), 'orderai.sqlite3')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# In-memory result store keyed by a one-time token
RESULT_STORE: Dict[str, Dict[str, bytes]] = {}
# Batch store maps a batch token to a list of item tokens
BATCH_STORE: Dict[str, List[str]] = {}
# Simple schedule store: token -> ISO date string
SCHEDULE_STORE: Dict[str, str] = {}
# Decisions store for Not have items: key(fish|pack|order) -> decision
DECISION_STORE: Dict[str, str] = {}


# ----- Database Functions -----
def init_database():
    """Initialize the SQLite database with required tables."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Create processing sessions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processing_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_token TEXT UNIQUE NOT NULL,
            batch_token TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            processing_type TEXT NOT NULL, -- 'single' or 'batch'
            total_items INTEGER DEFAULT 0,
            full_items INTEGER DEFAULT 0,
            not_full_items INTEGER DEFAULT 0,
            not_have_items INTEGER DEFAULT 0,
            total_kg REAL DEFAULT 0,
            full_kg REAL DEFAULT 0,
            not_full_kg REAL DEFAULT 0,
            not_have_kg REAL DEFAULT 0
        )
    ''')
    
    # Create uploaded files table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS uploaded_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER,
            file_type TEXT NOT NULL, -- 'stock' or 'order'
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id)
        )
    ''')
    
    # Create processing results table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processing_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER,
            order_file_id INTEGER,
            fish_name TEXT,
            packed_size TEXT,
            order_carton INTEGER,
            stock_carton INTEGER,
            order_kg_per_ctn REAL,
            stock_kg_per_ctn REAL,
            balance_stock_carton INTEGER,
            mc_to_give INTEGER,
            can_fulfill_carton INTEGER,
            shortfall INTEGER,
            status TEXT,
            required_kg REAL,
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id),
            FOREIGN KEY (order_file_id) REFERENCES uploaded_files (id)
        )
    ''')

    # Create scheduled orders table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scheduled_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            order_file_id INTEGER NOT NULL,
            scheduled_on TEXT NOT NULL,
            UNIQUE(session_id, order_file_id),
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id),
            FOREIGN KEY (order_file_id) REFERENCES uploaded_files (id)
        )
    ''')

    # Create fish decisions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS fish_decisions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            fish_name TEXT NOT NULL,
            packed_size TEXT NOT NULL,
            order_name TEXT NOT NULL,
            decision TEXT NOT NULL,
                          UNIQUE(session_id, fish_name, packed_size, order_name),
              FOREIGN KEY (session_id) REFERENCES processing_sessions (id)
         )
     ''')
     
    # Create file comparison history table
    cursor.execute('''
         CREATE TABLE IF NOT EXISTS file_comparison_history (
             id INTEGER PRIMARY KEY AUTOINCREMENT,
             session_id INTEGER,
             order_file_id INTEGER,
             original_token TEXT,
             batch_token TEXT,
             comparison_data TEXT,
             changes_applied TEXT,
             comparison_summary TEXT,
             bangkok_datetime TEXT,
             created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
             FOREIGN KEY (session_id) REFERENCES processing_sessions(id),
             FOREIGN KEY (order_file_id) REFERENCES uploaded_files(id)
         )
    ''')
    
    # Check if order_file_id column exists, if not add it (migration)
    try:
        cursor.execute("SELECT order_file_id FROM processing_results LIMIT 1")
    except sqlite3.OperationalError:
        # Column doesn't exist, add it
        cursor.execute("ALTER TABLE processing_results ADD COLUMN order_file_id INTEGER")
        print("Added order_file_id column to processing_results table")
    
    conn.commit()
    conn.close()

def save_processing_session(session_token: str, batch_token: str = None, processing_type: str = 'single', summary: dict = None) -> int:
    """Save a processing session to database and return session ID."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Bangkok time if available
    created_at = None
    try:
        if ZoneInfo:
            created_at = datetime.now(ZoneInfo('Asia/Bangkok')).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        created_at = None

    if created_at:
        cursor.execute('''
            INSERT INTO processing_sessions 
            (session_token, batch_token, processing_type, total_items, full_items, not_full_items, not_have_items, 
             total_kg, full_kg, not_full_kg, not_have_kg, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_token,
            batch_token,
            processing_type,
            summary.get('total_items', 0) if summary else 0,
            summary.get('full', 0) if summary else 0,
            summary.get('not_full', 0) if summary else 0,
            summary.get('not_have', 0) if summary else 0,
            summary.get('total_kg_all', 0) if summary else 0,
            summary.get('total_kg_full', 0) if summary else 0,
            summary.get('total_kg_not_full', 0) if summary else 0,
            summary.get('total_kg_not_have', 0) if summary else 0,
            created_at,
        ))
    else:
        cursor.execute('''
            INSERT INTO processing_sessions 
            (session_token, batch_token, processing_type, total_items, full_items, not_full_items, not_have_items, 
             total_kg, full_kg, not_full_kg, not_have_kg)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_token,
            batch_token,
            processing_type,
            summary.get('total_items', 0) if summary else 0,
            summary.get('full', 0) if summary else 0,
            summary.get('not_full', 0) if summary else 0,
            summary.get('not_have', 0) if summary else 0,
            summary.get('total_kg_all', 0) if summary else 0,
            summary.get('total_kg_full', 0) if summary else 0,
            summary.get('total_kg_not_full', 0) if summary else 0,
            summary.get('total_kg_not_have', 0) if summary else 0
        ))
    
    session_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return session_id

def upsert_schedule(session_id: int, order_file_id: int, scheduled_on: str) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO scheduled_orders (session_id, order_file_id, scheduled_on)
        VALUES (?, ?, ?)
        ON CONFLICT(session_id, order_file_id) DO UPDATE SET scheduled_on = excluded.scheduled_on
    ''', (session_id, order_file_id, scheduled_on))
    conn.commit()
    conn.close()

def save_file_comparison_history(session_id: int, order_file_id: int, original_token: str, 
                                batch_token: str, comparison_data: dict, changes_applied: list) -> None:
    """Save file comparison history to database with Bangkok timezone."""
    from datetime import datetime, timezone, timedelta
    
    # Get Bangkok timezone (+7 hours from UTC)
    bangkok_tz = timezone(timedelta(hours=7))
    bangkok_time = datetime.now(bangkok_tz)
    bangkok_datetime_str = bangkok_time.strftime('%Y-%m-%d %H:%M:%S %Z')
    
    # Prepare comparison summary
    summary = comparison_data.get('summary', {})
    comparison_summary = {
        'added': summary.get('added', 0),
        'modified': summary.get('modified', 0),
        'deleted': summary.get('deleted', 0),
        'unchanged': summary.get('unchanged', 0)
    }
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute(
        '''INSERT INTO file_comparison_history 
           (session_id, order_file_id, original_token, batch_token, comparison_data, 
            changes_applied, comparison_summary, bangkok_datetime)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        (
            session_id, order_file_id, original_token, batch_token,
            json.dumps(comparison_data), json.dumps(changes_applied), json.dumps(comparison_summary),
            bangkok_datetime_str
        )
    )
    conn.commit()
    conn.close()

def get_file_comparison_history(batch_token: str) -> list:
    """Get file comparison history for a batch."""
    session_id = get_session_id_by_batch_token(batch_token)
    if not session_id:
        return []
    
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT fch.*, uf.original_filename
        FROM file_comparison_history fch
        JOIN uploaded_files uf ON fch.order_file_id = uf.id
        WHERE fch.session_id = ?
        ORDER BY fch.created_at DESC
    ''', (session_id,))
    
    history = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return history

def delete_schedule(session_id: int, order_file_id: int) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM scheduled_orders WHERE session_id = ? AND order_file_id = ?', (session_id, order_file_id))
    conn.commit()
    conn.close()

def upsert_fish_decision(session_id: int, fish_name: str, packed_size: str, order_name: str, decision: str) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO fish_decisions (session_id, fish_name, packed_size, order_name, decision)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(session_id, fish_name, packed_size, order_name) DO UPDATE SET decision = excluded.decision
    ''', (session_id, fish_name, packed_size, order_name, decision))
    conn.commit()
    conn.close()

def save_uploaded_file(session_id: int, file_type: str, original_filename: str, file_storage) -> str:
    """Save uploaded file to disk and record in database."""
    # Generate unique filename
    file_ext = os.path.splitext(original_filename)[1]
    stored_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
    
    # Save file to disk
    file_storage.save(file_path)
    file_size = os.path.getsize(file_path)
    
    # Record in database
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO uploaded_files (session_id, file_type, original_filename, stored_filename, file_size)
        VALUES (?, ?, ?, ?, ?)
    ''', (session_id, file_type, original_filename, stored_filename, file_size))
    
    conn.commit()
    conn.close()
    
    return stored_filename

def save_processing_results(session_id: int, results: List[Dict[str, Any]], order_file_id: int = None):
    """Save processing results to database."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    for result in results:
        cursor.execute('''
            INSERT INTO processing_results 
            (session_id, order_file_id, fish_name, packed_size, order_carton, stock_carton, order_kg_per_ctn, 
             stock_kg_per_ctn, balance_stock_carton, mc_to_give, can_fulfill_carton, 
             shortfall, status, required_kg)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_id,
            order_file_id,
            result.get('fish name', ''),
            result.get('packed size', ''),
            result.get('order_carton', 0),
            result.get('stock_carton', 0),
            result.get('order_kg_per_ctn', 0),
            result.get('stock_kg_per_ctn', 0),
            result.get('balance_stock_carton', 0),
            result.get('mc_to_give', 0),
            result.get('can_fulfill_carton', 0),
            result.get('shortfall', 0),
            result.get('status', ''),
            result.get('required_kg', 0)
        ))
    
    conn.commit()
    conn.close()

def get_recent_sessions(limit: int = 10) -> List[Dict[str, Any]]:
    """Get recent processing sessions for display."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT ps.*, 
               COUNT(CASE WHEN uf.file_type = 'stock' THEN 1 END) as stock_files,
               COUNT(CASE WHEN uf.file_type = 'order' THEN 1 END) as order_files,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'stock' THEN uf.original_filename END) as stock_filename,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'order' THEN uf.original_filename END, ', ') as order_filenames
        FROM processing_sessions ps
        LEFT JOIN uploaded_files uf ON ps.id = uf.session_id
        GROUP BY ps.id
        ORDER BY datetime(ps.created_at) DESC
        LIMIT ?
    ''', (limit,))
    
    sessions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return sessions

def get_session_id_by_batch_token(batch_token: str) -> Optional[int]:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
    row = cursor.fetchone()
    conn.close()
    return int(row[0]) if row else None

def get_schedules_for_batch(batch_token: str) -> Dict[str, str]:
    """Return mapping of order original_filename -> scheduled_on for a batch token."""
    mapping: Dict[str, str] = {}
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT uf.original_filename AS order_name, so.scheduled_on
        FROM processing_sessions ps
        JOIN uploaded_files uf ON uf.session_id = ps.id AND uf.file_type = 'order'
        LEFT JOIN scheduled_orders so ON so.session_id = ps.id AND so.order_file_id = uf.id
        WHERE ps.batch_token = ?
    ''', (batch_token,))
    for row in cursor.fetchall():
        if row['scheduled_on']:
            mapping[str(row['order_name'])] = str(row['scheduled_on'])
    conn.close()
    return mapping

def get_session_by_id(session_id: int) -> Optional[Dict[str, Any]]:
    """Get a specific session with its files and results."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get session info
    cursor.execute('''
        SELECT ps.*, 
               COUNT(CASE WHEN uf.file_type = 'stock' THEN 1 END) as stock_files,
               COUNT(CASE WHEN uf.file_type = 'order' THEN 1 END) as order_files,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'stock' THEN uf.original_filename END) as stock_filename,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'order' THEN uf.original_filename END, ', ') as order_filenames
        FROM processing_sessions ps
        LEFT JOIN uploaded_files uf ON ps.id = uf.session_id
        WHERE ps.id = ?
        GROUP BY ps.id
    ''', (session_id,))
    
    session = cursor.fetchone()
    if not session:
        conn.close()
        return None
    
    session = dict(session)
    
    # Get processing results grouped by order file for batch sessions
    if session.get('processing_type') == 'batch':
        # Get all uploaded files for this session
        cursor.execute('''
            SELECT * FROM uploaded_files
            WHERE session_id = ? AND file_type = 'order'
            ORDER BY id
        ''', (session_id,))
        
        order_files = [dict(row) for row in cursor.fetchall()]
        session['order_files_list'] = order_files
        
        # Get results grouped by order file (we'll need to reconstruct this)
        cursor.execute('''
            SELECT * FROM processing_results
            WHERE session_id = ?
            ORDER BY id
        ''', (session_id,))
        
        all_results = [dict(row) for row in cursor.fetchall()]
        session['all_results'] = all_results
    else:
        # Get processing results for single sessions
        cursor.execute('''
            SELECT * FROM processing_results
            WHERE session_id = ?
            ORDER BY id
        ''', (session_id,))
        
        results = [dict(row) for row in cursor.fetchall()]
        session['results'] = results
    
    conn.close()
    return session

def restore_batch_session_to_memory(session: Dict[str, Any]) -> str:
    """Restore a batch session to memory stores and return batch_token."""
    if not session.get('batch_token'):
        return None
    
    batch_token = session['batch_token']
    stock_name = session.get('stock_filename', 'Unknown Stock')
    
    # Get order files and reprocess them from the actual stored files
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get stock file first
    cursor.execute('''
        SELECT stored_filename FROM uploaded_files
        WHERE session_id = ? AND file_type = 'stock'
        LIMIT 1
    ''', (session['id'],))
    stock_file_row = cursor.fetchone()
    
    if not stock_file_row:
        conn.close()
        return None
    
    # Load stock data from saved file
    stock_file_path = os.path.join(UPLOAD_FOLDER, stock_file_row['stored_filename'])
    try:
        with open(stock_file_path, 'rb') as f:
            stock_ds = load_excel(f)
    except:
        conn.close()
        return None
    
    # Get order files
    cursor.execute('''
        SELECT id, original_filename, stored_filename FROM uploaded_files
        WHERE session_id = ? AND file_type = 'order'
        ORDER BY id
    ''', (session['id'],))
    order_files = cursor.fetchall()
    
    conn.close()
    
    # Process each order file by reloading from the stored file
    files_data = {}
    for order_file in order_files:
        try:
            # Load the order file from disk (this will include any updates)
            order_file_path = os.path.join(UPLOAD_FOLDER, order_file['stored_filename'])
            with open(order_file_path, 'rb') as f:
                order_ds = load_excel(f)
            
            # Set the original filename for display purposes
            order_ds.name = order_file['original_filename']
            
            # Recompute matches using current stock data
            result_rows = compute_matches(stock_ds.rows, order_ds.rows)
            
            files_data[order_file['id']] = {
                'filename': order_file['original_filename'],
                'results': result_rows,
                'order_data': order_ds.rows  # Store original order data
            }
            
        except Exception as e:
            print(f"Warning: Failed to reload order file {order_file['original_filename']}: {e}")
            # Fallback to empty results
            files_data[order_file['id']] = {
                'filename': order_file['original_filename'],
                'results': [],
                'order_data': []
            }
    

    
    # Create tokens for each order file and populate RESULT_STORE
    token_list = []
    
    for file_id, file_data in files_data.items():
        token = uuid.uuid4().hex
        order_name = file_data['filename']
        result_rows = file_data['results']
        
        # Calculate summary for this order file
        def sum_required_kg(rows):
            return round(sum(float(r.get('required_kg', 0) or 0) for r in rows), 3)
        
        def sum_ready_kg(rows):
            return round(sum(float(r.get('stock_carton', 0) or 0) * float(r.get('stock_kg_per_ctn', 0) or 0) for r in rows), 3)
        
        summary = {
            'total_items': len(result_rows),
            'full': sum(1 for r in result_rows if r.get('status') == 'Full'),
            'not_full': sum(1 for r in result_rows if r.get('status') == 'Not Full'),
            'not_have': sum(1 for r in result_rows if r.get('status') == 'Not have'),
            'total_kg_all': sum_required_kg(result_rows),
            'total_kg_full': sum_required_kg([r for r in result_rows if r.get('status') == 'Full']),
            'total_kg_not_full': sum_required_kg([r for r in result_rows if r.get('status') == 'Not Full']),
            'total_kg_not_have': sum_required_kg([r for r in result_rows if r.get('status') == 'Not have']),
            'ready_kg': sum_ready_kg(result_rows)
        }
        
        # Create Excel/PDF bytes for historical data
        excel_bytes = rows_to_excel_bytes(result_rows)
        pdf_bytes = rows_to_pdf_bytes(result_rows)
        
        # Store in RESULT_STORE
        order_basename = os.path.splitext(order_name)[0]
        
        # Use the actual order data from the reloaded file
        order_data = file_data['order_data']
        original_order_data = []
        for row in order_data:
            normalized = try_map_row(row)
            weight_mc = normalized.get('weight_mc', '')
            
            # First try to parse order_kg_per_ctn from weight_mc
            try:
                order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
            except Exception:
                order_kg_per_ctn = parse_kg_per_carton(weight_mc)
            
            # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
            if not weight_mc and order_kg_per_ctn == 0:
                try:
                    order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                    if order_kg_per_ctn > 0:
                        weight_mc = str(order_kg_per_ctn)
                except:
                    pass
            elif not weight_mc and order_kg_per_ctn > 0:
                weight_mc = str(order_kg_per_ctn)
            
            order_row = {
                'fish name': normalized.get('fish name', ''),
                'packed size': normalized.get('packed size', ''),
                'pack': normalized.get('pack', ''),
                'total carton': to_int(normalized.get('total carton', 0)),
                'weight_mc': weight_mc,
                'order_kg_per_ctn': order_kg_per_ctn,
                'remark': normalized.get('remark', '')
            }
            original_order_data.append(order_row)
        
        RESULT_STORE[token] = {
            "excel": excel_bytes,
            "pdf": pdf_bytes,
            "excel_name": f"{order_basename} Calculation.xlsx",
            "pdf_name": f"{order_basename} Calculation.pdf",
            "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
            "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
            "stock_name": stock_name.encode("utf-8"),
            "order_name": order_name.encode("utf-8"),
            "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
        }
        
        token_list.append(token)
    
    # Store in BATCH_STORE
    BATCH_STORE[batch_token] = token_list
    
    return batch_token

def get_session_by_token(session_token: str) -> Optional[Dict[str, Any]]:
    """Get a session by its token."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('SELECT id FROM processing_sessions WHERE session_token = ?', (session_token,))
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return get_session_by_id(result['id'])
    return None

# Initialize database on startup
init_database()

# ----- Data Models -----
@dataclass
class Dataset:
    name: str
    rows: List[Dict[str, Any]]
    sheet_names: List[str]


# ----- Helpers -----
NORMALIZE_MAP = {
    "fish name": ["fish name", "fish", "product", "product name", "name"],
    "packed size": ["packed size", "pack", "pack size", "size"],
    "pack": ["pack", "packed size", "pack size", "size"],
    "total carton": [
        "total carton",
        "total_ctn",
        "total ctn",
        "ctn",
        "carton",
        "cartons",
        "qty",
        "quantity",
    ],
    "weight_mc": [
        "weight_mc",
        "weight mt",
        "weight_mt",
        "net_weigh",
        "net weight",
        "weight per mc",
        "mc_weight",
        "kg/ctn",
        "kg per ctn",
        "weight mc",
        "order kg/ctn",
    ],
    "remark": ["remark", "remarks", "note", "notes", "comment", "comments"],
}


def load_excel(file_storage, preferred_sheet: Optional[str] = None) -> Dataset:
    # Reset pointer and load workbook
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    wb = load_workbook(file_storage, data_only=True)
    sheet_name = preferred_sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        headers = []
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        record = {}
        empty = True
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val not in (None, ""):
                empty = False
            record[h] = val
        if not empty:
            rows.append(record)

    return Dataset(name=getattr(file_storage, "filename", "uploaded.xlsx"), rows=rows, sheet_names=wb.sheetnames)


def load_csv(file_storage) -> Dataset:
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:  # noqa: BLE001
        text = raw.decode("latin-1", errors="ignore")
    reader = csv.DictReader(text.splitlines())
    rows: List[Dict[str, Any]] = []
    for r in reader:
        # Drop None keys if any
        rows.append({(k or ""): v for k, v in r.items()})
    return Dataset(name=getattr(file_storage, "filename", "uploaded.csv"), rows=rows, sheet_names=["CSV"])


def load_tabular(file_storage) -> Dataset:
    filename = (getattr(file_storage, "filename", "") or "").lower()
    if filename.endswith(".csv"):
        return load_csv(file_storage)
    # default to excel
    return load_excel(file_storage)


def try_map_row(row: Dict[str, Any]) -> Dict[str, Any]:
    lower_map = {k.lower().strip(): k for k in row.keys()}
    normalized: Dict[str, Any] = {}
    for key, aliases in NORMALIZE_MAP.items():
        value = None
        for alias in aliases:
            if alias in lower_map:
                value = row.get(lower_map[alias])
                break
        normalized[key] = value
    return normalized


def normalize_text_val(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().upper().replace("  ", " ")


_PARENS_RE = re.compile(r"\([^\)]*\)")
_MULTISPACE_RE = re.compile(r"\s+")
_RANGE_RE = re.compile(r"(\d+)\s*[-–]\s*(\d+)")
_PUNCT_BREAK_RE = re.compile(r"[\./_,]+")
_NON_ALNUM_RE = re.compile(r"[^A-Z0-9]+")


def canonicalize_product(text: Any) -> str:
    s = normalize_text_val(text)
    s = _PARENS_RE.sub(" ", s)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    # remove unit/noise tokens
    stop = {
        "G", "GM", "GMS", "GRAM", "GRAMS", "PCS", "PC", "CTN", "CTNS", "GL", "GLAZE",
        "WITH", "PRINT", "BAG", "RIDER", "STICKER", "PACK", "SIZE", "KG",
        # common glaze percentages as numbers
        "5", "10", "15", "20", "25", "30", "35", "40",
    }
    tokens = [t for t in _MULTISPACE_RE.split(s) if t]
    filtered: List[str] = []
    for t in tokens:
        if t in stop:
            continue
        filtered.append(t)
    # Return a compact key without spaces so 'SILVER CARP' == 'SILVERCARP'
    return "".join(filtered)


def canonicalize_pack(text: Any) -> str:
    s = normalize_text_val(text)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    s = _MULTISPACE_RE.sub(" ", s)
    # Return compact form (no spaces) to ignore spacing differences
    return s.replace(" ", "").strip()


_MASS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG|KGS|G|GM|GRAM|GRAMS)")
_MULT_X_RE = re.compile(r"X\s*(\d+)")


def parse_kg_per_carton(text: Any) -> float:
    """Extract approximate kg per carton from a pack description, e.g. '1 KG X 10 BAG/CTN' -> 10.0.
    Returns 0.0 if cannot parse.
    """
    s = normalize_text_val(text)
    if not s:
        return 0.0
    s = s.replace("×", "X")
    # Combine first mass block with a nearby multiplier if present
    match = _MASS_RE.search(s)
    if not match:
        return 0.0
    value = float(match.group(1))
    unit = match.group(2)
    kg = value if unit.startswith("K") else value / 1000.0
    # Look ahead for a multiplier within the next ~20 chars
    tail = s[match.end():match.end() + 30]
    m2 = _MULT_X_RE.search(tail)
    mult = float(m2.group(1)) if m2 else 1.0
    return kg * mult


def to_int(val: Any) -> int:
    try:
        if val is None or val == "":
            return 0
        return int(float(val))
    except Exception:  # noqa: BLE001
        return 0


def compute_matches(stock_rows: List[Dict[str, Any]], order_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Build stock lookups
    stock_by_prod_pack: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_by_prod: Dict[str, Dict[str, Any]] = {}

    for r in stock_rows:
        nr = try_map_row(r)
        prod_key = canonicalize_product(nr.get("fish name"))
        pack_key = canonicalize_pack(nr.get("packed size"))
        qty = to_int(nr.get("total carton"))
        kg_per_ctn = parse_kg_per_carton(nr.get("packed size"))
        if not prod_key and not pack_key:
            continue
        if prod_key:
            agg = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
            agg["ctn"] += qty
            # prefer non-zero kg_per_ctn when available
            if agg.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
                agg["kg_per_ctn"] = kg_per_ctn
            stock_by_prod[prod_key] = agg
        key = (prod_key, pack_key)
        agg2 = stock_by_prod_pack.get(key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
        agg2["ctn"] += qty
        if agg2.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
            agg2["kg_per_ctn"] = kg_per_ctn
        stock_by_prod_pack[key] = agg2

    results: List[Dict[str, Any]] = []
    for r in order_rows:
        nr = try_map_row(r)
        fish_text = nr.get("fish name")
        pack_text = nr.get("packed size")
        prod_key = canonicalize_product(fish_text)
        pack_key = canonicalize_pack(pack_text)
        order_qty = to_int(nr.get("total carton"))
        weight_mc = nr.get("weight_mc")
        try:
            order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
        except Exception:  # noqa: BLE001
            order_kg_per_ctn = parse_kg_per_carton(weight_mc)

        matched_by = ""
        stock_qty = 0
        # Prefer exact product+pack match when pack is present
        if pack_key and (prod_key, pack_key) in stock_by_prod_pack:
            stock_info = stock_by_prod_pack.get((prod_key, pack_key), {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product+pack"
        elif prod_key in stock_by_prod:
            stock_info = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product"
        else:
            stock_pack_kg = 0.0

        # Weight-aware calculations
        effective_stock_ctn = stock_qty
        mc_to_give = 0
        if order_kg_per_ctn and stock_pack_kg:
            total_stock_kg = stock_qty * stock_pack_kg
            # how many order-equivalent cartons can stock satisfy
            effective_stock_ctn = int(total_stock_kg // max(order_kg_per_ctn, 0.0001))
            # MC to pick from stock to satisfy order requirement
            required_kg = order_qty * order_kg_per_ctn
            mc_needed = math.ceil(required_kg / max(stock_pack_kg, 0.0001))
            mc_to_give = min(stock_qty, mc_needed)
        else:
            mc_to_give = min(stock_qty, order_qty)

        if effective_stock_ctn <= 0:
            status = "Not have"
        elif effective_stock_ctn < order_qty:
            status = "Not Full"
        else:
            status = "Full"

        fulfilled_ctn = min(order_qty, effective_stock_ctn)
        # Balance stock after giving the computed MC from stock
        balance_after_order = max(stock_qty - mc_to_give, 0)

        result = {
            "fish name": fish_text,
            "packed size": pack_text,
            "order_carton": order_qty,
            "stock_carton": stock_qty,  # raw MC from stock file
            "can_fulfill_carton": fulfilled_ctn,
            "shortfall": max(order_qty - effective_stock_ctn, 0),
            "status": status,
            "matched_by": matched_by,
            "order_kg_per_ctn": round(order_kg_per_ctn, 3) if order_kg_per_ctn else 0,
            "stock_kg_per_ctn": round(stock_pack_kg, 3) if stock_pack_kg else 0,
            "balance_stock_carton": balance_after_order,
            "mc_to_give": mc_to_give,
            "required_kg": round(order_qty * (order_kg_per_ctn or 0), 3),
        }
        results.append(result)

    return results


def rows_to_excel_bytes(rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()

    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Status",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
    ]

    def write_sheet(ws, data_rows: List[Dict[str, Any]]):
        ws.append(headers)
        for r in data_rows:
            ws.append([
                r.get("fish name", ""),
                r.get("packed size", ""),
                r.get("order_carton", 0),
                r.get("stock_carton", 0),
                r.get("can_fulfill_carton", 0),
                r.get("shortfall", 0),
                r.get("status", ""),
                r.get("order_kg_per_ctn", 0),
                r.get("stock_kg_per_ctn", 0),
                r.get("balance_stock_carton", 0),
            ])
        ws.freeze_panes = "A2"
        widths = [35, 18, 12, 12, 14, 12, 12, 14, 14, 18]
        for idx, width in enumerate(widths, start=1):
            col = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col].width = width

    # Create sheets: All, Full, Not Full, Not have
    ws_all = wb.active
    ws_all.title = "All"
    write_sheet(ws_all, rows)

    ws_full = wb.create_sheet(title="Full")
    write_sheet(ws_full, [r for r in rows if r.get("status") == "Full"])

    ws_nf = wb.create_sheet(title="Not Full")
    write_sheet(ws_nf, [r for r in rows if r.get("status") == "Not Full"])

    ws_nh = wb.create_sheet(title="Not have")
    write_sheet(ws_nh, [r for r in rows if r.get("status") == "Not have"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def rows_to_pdf_bytes(rows: List[Dict[str, Any]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Order Availability Result", styles["Heading2"])
    story.append(title)
    story.append(Spacer(1, 10))

    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Status",
    ]
    data = [headers]
    for r in rows:
        data.append([
            r.get("fish name", ""),
            r.get("packed size", ""),
            r.get("order_carton", 0),
            r.get("stock_carton", 0),
            r.get("order_kg_per_ctn", 0),
            r.get("stock_kg_per_ctn", 0),
            r.get("balance_stock_carton", 0),
            r.get("can_fulfill_carton", 0),
            r.get("shortfall", 0),
            r.get("status", ""),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fcfcfc")]),
    ]))
    story.append(table)

    doc.build(story)
    return buffer.getvalue()


# ----- Routes -----
@app.get("/")
def index():
    recent_sessions = get_recent_sessions(limit=5)
    return render_template("index.html", recent_sessions=recent_sessions)


@app.post('/delete-session')
def delete_session():
    session_id = request.form.get('session_id', type=int)
    if not session_id:
        flash('Invalid session selected.', 'error')
        return redirect(url_for('index'))
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        # If batch, capture batch_token to also clear in-memory batch if present
        cursor.execute('SELECT batch_token FROM processing_sessions WHERE id=?', (session_id,))
        row = cursor.fetchone()
        batch_token = row[0] if row else None

        # Delete related rows
        cursor.execute('DELETE FROM scheduled_orders WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM fish_decisions WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM processing_results WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM uploaded_files WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM processing_sessions WHERE id=?', (session_id,))
        conn.commit()
        conn.close()

        # Clear in-memory stores for that batch, if exists
        if batch_token and batch_token in BATCH_STORE:
            for tok in BATCH_STORE.get(batch_token, []):
                RESULT_STORE.pop(tok, None)
            BATCH_STORE.pop(batch_token, None)
        flash('History entry deleted.', 'success')
    except Exception as e:
        flash(f'Failed to delete: {e}', 'error')
    return redirect(url_for('index'))


@app.post("/process")
def process():
    stock_file = request.files.get("stock_file")
    order_file = request.files.get("order_file")

    if not stock_file or not order_file:
        flash("Please upload both Stock and Order Excel files.", "error")
        return redirect(url_for("index"))

    try:
        stock_ds = load_tabular(stock_file)
        order_ds = load_tabular(order_file)
        result_rows = compute_matches(stock_ds.rows, order_ds.rows)
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to process files: {exc}", "error")
        return redirect(url_for("index"))

    def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
        total = 0.0
        for r in rows:
            try:
                total += float(r.get("required_kg", 0) or 0)
            except Exception:  # noqa: BLE001
                pass
        return round(total, 3)

    summary = {
        "total_items": int(len(result_rows)),
        "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
        "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
        "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
        "total_kg_all": sum_required_kg(result_rows),
        "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
        "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
        "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
    }

    # Save to database
    try:
        # Save processing session
        session_id = save_processing_session(
            session_token=uuid.uuid4().hex,
            processing_type='single',
            summary=summary
        )
        
        # Save uploaded files
        stock_file.seek(0)  # Reset file pointer
        save_uploaded_file(session_id, 'stock', stock_file.filename or 'stock.xlsx', stock_file)
        
        order_file.seek(0)  # Reset file pointer
        save_uploaded_file(session_id, 'order', order_file.filename or 'order.xlsx', order_file)
        
        # Save processing results
        save_processing_results(session_id, result_rows)
        
    except Exception as e:
        print(f"Warning: Failed to save to database: {e}")

    # store result in memory for download via token
    excel_bytes = rows_to_excel_bytes(result_rows)
    pdf_bytes = rows_to_pdf_bytes(result_rows)
    token = uuid.uuid4().hex
    order_basename = os.path.splitext(order_ds.name or "order")[0]
    
    # Store original order data for later editing
    original_order_data = []
    for row in order_ds.rows:
        normalized = try_map_row(row)
        weight_mc = normalized.get('weight_mc', '')
        
        # First try to parse order_kg_per_ctn from weight_mc
        try:
            order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
        except Exception:
            order_kg_per_ctn = parse_kg_per_carton(weight_mc)
        
        # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
        if not weight_mc and order_kg_per_ctn == 0:
            try:
                order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                if order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
            except:
                pass
        elif not weight_mc and order_kg_per_ctn > 0:
            weight_mc = str(order_kg_per_ctn)
        
        order_row = {
            'fish name': normalized.get('fish name', ''),
            'packed size': normalized.get('packed size', ''),
            'pack': normalized.get('pack', ''),
            'total carton': to_int(normalized.get('total carton', 0)),
            'weight_mc': weight_mc,
            'order_kg_per_ctn': order_kg_per_ctn,
            'remark': normalized.get('remark', '')
        }
        original_order_data.append(order_row)
    
    RESULT_STORE[token] = {
        "excel": excel_bytes,
        "pdf": pdf_bytes,
        "excel_name": f"{order_basename} Calculation.xlsx",
        "pdf_name": f"{order_basename} Calculation.pdf",
        # Non-bytes metadata for rendering view routes
        "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
        "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
        "stock_name": stock_ds.name.encode("utf-8"),
        "order_name": order_ds.name.encode("utf-8"),
        "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
    }

    table_records = result_rows
    return render_template(
        "result.html",
        summary=summary,
        records=table_records,
        stock_name=stock_ds.name,
        order_name=order_ds.name,
        download_token=token,
    )


@app.get("/batch")
def batch_index():
    return render_template("batch.html")


@app.post("/process-batch")
def process_batch():
    stock_file = request.files.get("stock_file")
    order_files = request.files.getlist("order_files")
    if not stock_file or not order_files:
        flash("Please upload one Stock file and up to 32 Order files.", "error")
        return redirect(url_for("batch_index"))
    if len(order_files) > 32:
        flash("You can upload at most 32 order files.", "error")
        return redirect(url_for("batch_index"))

    try:
        stock_ds = load_tabular(stock_file)
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to read stock file: {exc}", "error")
        return redirect(url_for("batch_index"))

    batch_token = uuid.uuid4().hex
    token_list: List[str] = []

    # Save batch session to database
    batch_session_id = None
    try:
        batch_session_id = save_processing_session(
            session_token=uuid.uuid4().hex,
            batch_token=batch_token,
            processing_type='batch'
        )
        
        # Save stock file for batch
        stock_file.seek(0)
        save_uploaded_file(batch_session_id, 'stock', stock_file.filename or 'stock.xlsx', stock_file)
    except Exception as e:
        print(f"Warning: Failed to save batch session: {e}")

    # Accumulate unique Not have fish across all orders in this batch
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for of in order_files:
        try:
            order_ds = load_tabular(of)
            result_rows = compute_matches(stock_ds.rows, order_ds.rows)
            # Build summary like in single process
            def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("required_kg", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            # Compute ready kg from stock present for this order (sum over rows of stock_ctn * stock_kg_per_ctn)
            def sum_ready_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("stock_carton", 0) or 0) * float(r.get("stock_kg_per_ctn", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            summary = {
                "total_items": int(len(result_rows)),
                "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
                "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
                "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
                "total_kg_all": sum_required_kg(result_rows),
                "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
                "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
                "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
                "ready_kg": sum_ready_kg(result_rows),
            }

            # Build per-order Not have fish, aggregated by fish+pack with needed kg
            for row in result_rows:
                if row.get("status") == "Not have":
                    fish_name = str(row.get("fish name", ""))
                    pack_text = str(row.get("packed size", ""))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)
                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    # attach decision if already set
                    decision_key = f"{fish_name}|{pack_text}|{order_ds.name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_ds.name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })

            # Save individual order processing to database
            try:
                if batch_session_id:
                    # Save order file and get its ID
                    of.seek(0)
                    order_file_stored_name = save_uploaded_file(batch_session_id, 'order', of.filename or 'order.xlsx', of)
                    
                    # Get the file ID for linking results
                    conn = sqlite3.connect(DATABASE_PATH)
                    cursor = conn.cursor()
                    cursor.execute('SELECT id FROM uploaded_files WHERE session_id = ? AND stored_filename = ?', 
                                 (batch_session_id, order_file_stored_name))
                    order_file_record = cursor.fetchone()
                    order_file_id = order_file_record[0] if order_file_record else None
                    conn.close()
                    
                    # Save processing results linked to this order file
                    save_processing_results(batch_session_id, result_rows, order_file_id)
                    # Persist any existing in-memory schedule for this token
                    # Note: New sessions won't have this yet; restored sessions will
                    scheduled_on = SCHEDULE_STORE.get(token)
                    if scheduled_on and order_file_id:
                        upsert_schedule(batch_session_id, order_file_id, scheduled_on)
            except Exception as e:
                print(f"Warning: Failed to save order processing: {e}")

            excel_bytes = rows_to_excel_bytes(result_rows)
            pdf_bytes = rows_to_pdf_bytes(result_rows)
            token = uuid.uuid4().hex
            order_basename = os.path.splitext(order_ds.name or "order")[0]
            # Store original order data for later editing
            original_order_data = []
            for row in order_ds.rows:
                normalized = try_map_row(row)
                weight_mc = normalized.get('weight_mc', '')
                
                # First try to parse order_kg_per_ctn from weight_mc
                try:
                    order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
                except Exception:
                    order_kg_per_ctn = parse_kg_per_carton(weight_mc)
                
                # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
                if not weight_mc and order_kg_per_ctn == 0:
                    try:
                        order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                        if order_kg_per_ctn > 0:
                            weight_mc = str(order_kg_per_ctn)
                    except:
                        pass
                elif not weight_mc and order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
                
                order_row = {
                    'fish name': normalized.get('fish name', ''),
                    'packed size': normalized.get('packed size', ''),
                    'pack': normalized.get('pack', ''),
                    'total carton': to_int(normalized.get('total carton', 0)),
                    'weight_mc': weight_mc,
                    'order_kg_per_ctn': order_kg_per_ctn,
                    'remark': normalized.get('remark', '')
                }
                original_order_data.append(order_row)

            RESULT_STORE[token] = {
                "excel": excel_bytes,
                "pdf": pdf_bytes,
                "excel_name": f"{order_basename} Calculation.xlsx",
                "pdf_name": f"{order_basename} Calculation.pdf",
                "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
                "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
                "stock_name": stock_ds.name.encode("utf-8"),
                "order_name": order_ds.name.encode("utf-8"),
                "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
            }
            token_list.append(token)
        except Exception as exc:  # noqa: BLE001
            flash(f"Failed to process order file {getattr(of,'filename','unknown')}: {exc}", "error")

    BATCH_STORE[batch_token] = token_list

    # Load persisted schedules from DB and decorate items
    try:
        schedule_map = get_schedules_for_batch(batch_token)
    except Exception:
        schedule_map = {}

    # Build summary items
    items = []
    for t in token_list:
        entry = RESULT_STORE.get(t, {})
        try:
            # Decode stored summary/metadata
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            # If ready_kg missing (older entry), compute from rows
            if "ready_kg" not in summary:
                try:
                    rows_tmp = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                    ready = 0.0
                    for rr in rows_tmp:
                        ready += float(rr.get("stock_carton", 0) or 0) * float(rr.get("stock_kg_per_ctn", 0) or 0)
                    summary["ready_kg"] = round(ready, 3)
                except Exception:
                    summary["ready_kg"] = 0.0

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": schedule_map.get(order_name) or SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    # Build chart data
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]

    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    # Turn fish groups dict into a list for template
    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    # Recommendations: sort orders by number of Full items desc; include ready_kg
    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "full_kg": it["summary"].get("total_kg_full", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["full_kg"]),
        reverse=True,
    )

    # Persist aggregated batch summary numbers for accurate history display
    try:
        if batch_session_id:
            agg_total_items = sum(it["summary"].get("total_items", 0) for it in items)
            agg_full = sum(it["summary"].get("full", 0) for it in items)
            agg_nf = sum(it["summary"].get("not_full", 0) for it in items)
            agg_nh = sum(it["summary"].get("not_have", 0) for it in items)
            agg_kg_all = sum(it["summary"].get("total_kg_all", 0) for it in items)
            agg_kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
            agg_kg_nf = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
            agg_kg_nh = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute(
                '''UPDATE processing_sessions
                   SET total_items=?, full_items=?, not_full_items=?, not_have_items=?,
                       total_kg=?, full_kg=?, not_full_kg=?, not_have_kg=?
                 WHERE id=?''',
                (
                    int(agg_total_items), int(agg_full), int(agg_nf), int(agg_nh),
                    float(agg_kg_all), float(agg_kg_full), float(agg_kg_nf), float(agg_kg_nh),
                    batch_session_id,
                )
            )
            conn.commit()
            conn.close()
    except Exception:
        pass

    return redirect(url_for('view_batch', batch_token=batch_token))


@app.get("/get-comparison-history/<batch_token>")
def get_comparison_history(batch_token: str):
    """Get file comparison history for a batch."""
    try:
        history = get_file_comparison_history(batch_token)
        return {"success": True, "history": history}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.post("/clear-comparison-history/<batch_token>")
def clear_comparison_history(batch_token: str):
    """Clear comparison history for a batch (for testing/debugging)."""
    try:
        session_id = get_session_id_by_batch_token(batch_token)
        if session_id:
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM file_comparison_history WHERE session_id = ?', (session_id,))
            conn.commit()
            conn.close()
            return {"success": True, "message": "History cleared"}
        else:
            return {"success": False, "error": "Session not found"}, 404
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.get("/batch/<batch_token>")
def view_batch(batch_token: str):
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    items = []
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore") or stock_name
            # aggregate Not have by fish+pack with needed kg per order
            try:
                rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                for row in rows:
                    if row.get("status") == "Not have":
                        fish_name = str(row.get("fish name", ""))
                        pack_text = str(row.get("packed size", ""))
                        needed_kg = float(row.get("required_kg", 0) or 0)
                        key = (fish_name, pack_text)
                        group = fish_groups.get(key)
                        if not group:
                            group = {
                                "fish_name": fish_name,
                                "packed_size": pack_text,
                                "total_needed_kg": 0.0,
                                "orders": [],
                            }
                            fish_groups[key] = group
                        group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                        decision_key = f"{fish_name}|{pack_text}|{order_name}"
                        decision = DECISION_STORE.get(decision_key)
                        group["orders"].append({
                            "order_name": order_name,
                            "needed_kg": round(needed_kg, 3),
                            "decision": decision,
                        })
            except Exception:
                pass

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": get_schedules_for_batch(batch_token).get(order_name) or SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)
    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "full_kg": it["summary"].get("total_kg_full", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["full_kg"]),
        reverse=True,
    )

    return render_template(
        "summary.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups_list,
        fish_total_kg=fish_total_kg,
        recommendations=recommendations,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
        calendar_events=json.dumps(events),
    )


@app.post("/set-decision")
def set_decision():
    fish_name = request.form.get('fish_name') or ''
    packed_size = request.form.get('packed_size') or ''
    order_name = request.form.get('order_name') or ''
    decision = request.form.get('decision') or ''
    batch_token = request.form.get('batch_token') or ''
    redirect_to = request.form.get('redirect_to') or ''
    key = f"{fish_name}|{packed_size}|{order_name}"
    if decision:
        DECISION_STORE[key] = decision
        # Persist to DB (best-effort): attach to most recent batch session
        try:
            conn = sqlite3.connect(DATABASE_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM processing_sessions WHERE processing_type='batch' ORDER BY created_at DESC LIMIT 1")
            row = cursor.fetchone()
            if row:
                upsert_fish_decision(row['id'], fish_name, packed_size, order_name, decision)
            conn.close()
        except Exception:
            pass
    else:
        DECISION_STORE.pop(key, None)
    if batch_token:
        if redirect_to == 'fish_buy':
            return redirect(url_for('fish_buy', batch_token=batch_token))
        # default: bring user back to summary with fish tab active (handled by JS init)
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


def build_fish_groups_from_batch(batch_token: str) -> List[Dict[str, Any]]:
    tokens = BATCH_STORE.get(batch_token, [])
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
            for row in rows:
                if row.get("status") == "Not have":
                    fish_name = str(row.get("fish name", ""))
                    pack_text = str(row.get("packed size", ""))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)
                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    decision_key = f"{fish_name}|{pack_text}|{order_name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })
        except Exception:
            continue
    return [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]


@app.get("/fish-buy/<batch_token>")
def fish_buy(batch_token: str):
    # reuse batch view data
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    # stock name from first token
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        if stock_name:
            break

    items = []
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            items.append({"token": t, "order_name": order_name, "summary": summary})
        except Exception:
            continue

    fish_groups = build_fish_groups_from_batch(batch_token)

    # charts data (optional on this page)
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    return render_template(
        "fish_buy.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
    )


def fish_groups_to_excel_bytes(fish_groups: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fish Summary"
    ws.append(["Fish Name", "Packed Size", "Total Needed KG", "Orders Count"])
    for g in fish_groups:
        ws.append([g.get("fish_name", ""), g.get("packed_size", ""), g.get("total_needed_kg", 0), len(g.get("orders", []))])
    ws.freeze_panes = "A2"

    # Decisions sheet: only orders with a decision
    ws2 = wb.create_sheet("Decisions")
    ws2.append(["Fish Name", "Packed Size", "Order File", "Needed KG", "Decision"])
    for g in fish_groups:
        for o in g.get("orders", []):
            if o.get("decision"):
                ws2.append([g.get("fish_name", ""), g.get("packed_size", ""), o.get("order_name", ""), o.get("needed_kg", 0), o.get("decision")])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/download/fish-excel")
def download_fish_excel():
    batch_token = request.args.get('batch', type=str)
    if not batch_token or batch_token not in BATCH_STORE:
        flash("Unknown batch.", "error")
        return redirect(url_for('batch_index'))
    fish_groups = build_fish_groups_from_batch(batch_token)
    raw = fish_groups_to_excel_bytes(fish_groups)
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Fish Decisions - {batch_token}.xlsx",
    )


@app.post("/schedule-order")
def schedule_order():
    token = request.form.get('token')
    date = request.form.get('date')
    batch_token = request.form.get('batch_token')
    if token and date:
        SCHEDULE_STORE[token] = date
        # Persist to DB if this token belongs to a batch restored session
        # Find the order name from RESULT_STORE
        entry = RESULT_STORE.get(token)
        if entry and batch_token:
            order_name = (entry.get('order_name') or b'').decode('utf-8', errors='ignore')
            stock_name = (entry.get('stock_name') or b'').decode('utf-8', errors='ignore')
            # Locate a session that matches this stock file and batch (best-effort)
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                # Find exact session by batch_token
                cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
                row = cursor.fetchone()
                if row:
                    session_id = row['id']
                    # Find order file id by original filename
                    cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", (session_id, order_name))
                    of = cursor.fetchone()
                    if of:
                        upsert_schedule(session_id, of['id'], date)
                conn.close()
            except Exception:
                pass
    if batch_token:
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


@app.post("/unschedule-order")
def unschedule_order():
    token = request.form.get('token')
    batch_token = request.form.get('batch_token')
    if token:
        SCHEDULE_STORE.pop(token, None)
        # Remove from DB as well (best-effort)
        entry = RESULT_STORE.get(token)
        if entry and batch_token:
            order_name = (entry.get('order_name') or b'').decode('utf-8', errors='ignore')
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
                row = cursor.fetchone()
                if row:
                    session_id = row['id']
                    cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", (session_id, order_name))
                    of = cursor.fetchone()
                    if of:
                        delete_schedule(session_id, of['id'])
                conn.close()
            except Exception:
                pass
    # For fetch usage, return a simple OK
    return "OK"


@app.post('/remove-order')
def remove_order():
    token = request.form.get('token')
    batch_token = request.form.get('batch_token')
    if not token or not batch_token:
        flash('Invalid remove request.', 'error')
        return redirect(url_for('batch_index'))

    # Capture entry (order name) before mutating in-memory stores
    entry_snapshot = RESULT_STORE.get(token)
    order_name_snapshot = (entry_snapshot.get('order_name') or b'').decode('utf-8', errors='ignore') if entry_snapshot else None

    # Remove from in-memory batch store
    tokens = BATCH_STORE.get(batch_token, [])
    if token in tokens:
        tokens.remove(token)
        BATCH_STORE[batch_token] = tokens
    # Also drop in-memory result for this token
    RESULT_STORE.pop(token, None)

    # Best-effort: remove the order file record and its results from DB for this batch
    try:
        order_name = order_name_snapshot
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1', (batch_token,))
        row = cursor.fetchone()
        if row and order_name:
            session_id = row['id']
            cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=?", (session_id, order_name))
            of = cursor.fetchone()
            if of:
                order_file_id = of['id']
                cursor.execute('DELETE FROM processing_results WHERE session_id=? AND order_file_id=?', (session_id, order_file_id))
                cursor.execute('DELETE FROM scheduled_orders WHERE session_id=? AND order_file_id=?', (session_id, order_file_id))
                cursor.execute('DELETE FROM uploaded_files WHERE id=?', (order_file_id,))
        conn.commit()
        conn.close()
    except Exception:
        pass

    flash('Order removed.', 'success')
    return redirect(url_for('view_batch', batch_token=batch_token))


@app.get("/result/<token>")
def view_result(token: str):
    entry = RESULT_STORE.get(token)
    if not entry:
        flash("Unknown or expired result token.", "error")
        return redirect(url_for("index"))
    try:
        rows = eval(entry.get("rows_json", b"[]"))  # noqa: S307
        summary = eval(entry.get("summary_json", b"{}"))  # noqa: S307
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
    except Exception:
        flash("Failed to load stored result.", "error")
        return redirect(url_for("index"))

    return render_template(
        "result.html",
        summary=summary,
        records=rows,
        stock_name=stock_name,
        order_name=order_name,
        download_token=token,
    )


@app.get("/download/excel")
def download_excel():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("excel") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("excel_name", "order_stock_result.xlsx")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=name,
    )


@app.get("/download/pdf")
def download_pdf():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("pdf") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("pdf_name", "order_stock_result.pdf")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=name,
    )


@app.get("/session/<int:session_id>")
def view_session(session_id: int):
    """View a saved processing session."""
    session = get_session_by_id(session_id)
    if not session:
        flash("Session not found.", "error")
        return redirect(url_for("index"))
    
    # For batch sessions, restore to memory and redirect to batch summary
    if session.get('processing_type') == 'batch':
        batch_token = restore_batch_session_to_memory(session)
        if batch_token:
            return redirect(url_for('view_batch', batch_token=batch_token))
        else:
            flash("Failed to restore batch session.", "error")
            return redirect(url_for("index"))
    
    # Handle single sessions
    results = session.get('results', [])
    records = []
    for r in results:
        record = {
            'fish name': r.get('fish_name', ''),
            'packed size': r.get('packed_size', ''),
            'order_carton': r.get('order_carton', 0),
            'stock_carton': r.get('stock_carton', 0),
            'order_kg_per_ctn': r.get('order_kg_per_ctn', 0),
            'stock_kg_per_ctn': r.get('stock_kg_per_ctn', 0),
            'balance_stock_carton': r.get('balance_stock_carton', 0),
            'mc_to_give': r.get('mc_to_give', 0),
            'can_fulfill_carton': r.get('can_fulfill_carton', 0),
            'shortfall': r.get('shortfall', 0),
            'status': r.get('status', ''),
            'required_kg': r.get('required_kg', 0)
        }
        records.append(record)
    
    # Create summary from session data
    summary = {
        'total_items': session.get('total_items', 0),
        'full': session.get('full_items', 0),
        'not_full': session.get('not_full_items', 0),
        'not_have': session.get('not_have_items', 0),
        'total_kg_all': session.get('total_kg', 0),
        'total_kg_full': session.get('full_kg', 0),
        'total_kg_not_full': session.get('not_full_kg', 0),
        'total_kg_not_have': session.get('not_have_kg', 0)
    }
    
    return render_template(
        "result.html",
        summary=summary,
        records=records,
        stock_name=session.get('stock_filename', 'Unknown Stock File'),
        order_name=session.get('order_filenames', 'Unknown Order File'),
        download_token=None,  # No download available for historical data
        is_historical=True
    )


@app.post("/compare-order-files")
def compare_order_files():
    """Compare an original order file with an updated version and return differences."""
    try:
        original_token = request.form.get('original_token')
        batch_token = request.form.get('batch_token')
        updated_file = request.files.get('updated_file')
        
        if not original_token or not updated_file:
            return {"success": False, "error": "Missing required parameters"}, 400
        
        # Get original file data from RESULT_STORE
        original_entry = RESULT_STORE.get(original_token)
        if not original_entry:
            return {"success": False, "error": "Original file not found"}, 404
        
        # Get original filename for validation
        try:
            original_filename = (original_entry.get("order_name") or b"").decode("utf-8", errors="ignore")
        except:
            original_filename = ""
        
        # Validate filename matches (base name without extension)
        if original_filename and updated_file.filename:
            import os
            original_base = os.path.splitext(original_filename)[0].lower()
            updated_base = os.path.splitext(updated_file.filename)[0].lower()
            
            if original_base != updated_base:
                return {
                    "success": False, 
                    "error": f"File name mismatch. Expected: {original_filename}, Got: {updated_file.filename}"
                }, 400
        
        # Parse original file data - use stored original order data if available
        try:
            if "original_order_json" in original_entry:
                # Use stored original order data
                original_rows = eval(original_entry.get("original_order_json", b"[]"))
            else:
                # Fallback: reconstruct from processed results
                processed_rows = eval(original_entry.get("rows_json", b"[]"))
                original_rows = []
                for row in processed_rows:
                    order_row = {
                        'fish name': row.get('fish name', ''),
                        'packed size': row.get('packed size', ''),
                        'total carton': row.get('order_carton', 0),
                        'weight_mc': '',
                        'order_kg_per_ctn': row.get('order_kg_per_ctn', 0)
                    }
                    original_rows.append(order_row)
        except:
            return {"success": False, "error": "Could not parse original file data"}, 500
        
        # Load and parse updated file
        try:
            updated_ds = load_tabular(updated_file)
            updated_rows = []
            for row in updated_ds.rows:
                normalized = try_map_row(row)
                weight_mc = normalized.get('weight_mc', '')
                
                # Parse order_kg_per_ctn from weight_mc
                try:
                    order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
                except Exception:
                    order_kg_per_ctn = parse_kg_per_carton(weight_mc)
                
                # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
                if not weight_mc and order_kg_per_ctn == 0:
                    try:
                        order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                        if order_kg_per_ctn > 0:
                            weight_mc = str(order_kg_per_ctn)
                    except:
                        pass
                elif not weight_mc and order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
                
                updated_rows.append({
                    'fish name': normalized.get('fish name', ''),
                    'packed size': normalized.get('packed size', ''),
                    'pack': normalized.get('pack', ''),
                    'total carton': to_int(normalized.get('total carton', 0)),
                    'weight_mc': weight_mc,
                    'order_kg_per_ctn': order_kg_per_ctn,
                    'remark': normalized.get('remark', '')
                })
        except Exception as e:
            return {"success": False, "error": f"Could not parse updated file: {str(e)}"}, 400
        
        # Perform comparison
        comparison_result = compare_order_file_data(original_rows, updated_rows)
        
        # Save comparison history to database
        try:
            session_id = get_session_id_by_batch_token(batch_token)
            if session_id:
                # Find order file ID
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", 
                             (session_id, original_filename))
                order_file_record = cursor.fetchone()
                
                if order_file_record:
                    order_file_id = order_file_record[0]
                    save_file_comparison_history(
                        session_id=session_id,
                        order_file_id=order_file_id,
                        original_token=original_token,
                        batch_token=batch_token,
                        comparison_data=comparison_result,
                        changes_applied=[]  # Will be filled when changes are applied
                    )
                
                conn.close()
        except Exception as e:
            print(f"Warning: Failed to save comparison history: {e}")
        
        return {
            "success": True,
            "comparison": comparison_result
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


def compare_order_file_data(original_rows, updated_rows):
    """Compare two sets of order data and return detailed differences."""
    # Create lookup dictionaries for comparison
    original_lookup = {}
    updated_lookup = {}
    
    # Build lookup for original data
    for row in original_rows:
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        original_lookup[key] = {
            'fish_name': row.get('fish name', ''),
            'packed_size': row.get('packed size', ''),
            'pack': row.get('pack', ''),
            'quantity': row.get('total carton', 0) if 'total carton' in row else row.get('order_carton', 0),
            'weight_mc': row.get('weight_mc', ''),
            'order_kg_per_ctn': row.get('order_kg_per_ctn', 0),
            'remark': row.get('remark', ''),
            'raw_data': row
        }
    
    # Build lookup for updated data
    for row in updated_rows:
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        updated_lookup[key] = {
            'fish_name': row.get('fish name', ''),
            'packed_size': row.get('packed size', ''),
            'pack': row.get('pack', ''),
            'quantity': row.get('total carton', 0),
            'weight_mc': row.get('weight_mc', ''),
            'order_kg_per_ctn': row.get('order_kg_per_ctn', 0),
            'remark': row.get('remark', ''),
            'raw_data': row
        }
    
    # Find changes
    changes = []
    all_keys = set(original_lookup.keys()) | set(updated_lookup.keys())
    
    added_count = 0
    modified_count = 0
    deleted_count = 0
    unchanged_count = 0
    
    for key in all_keys:
        original_item = original_lookup.get(key)
        updated_item = updated_lookup.get(key)
        
        if original_item and updated_item:
            # Item exists in both - check for modifications
            quantity_changed = original_item['quantity'] != updated_item['quantity']
            weight_changed = str(original_item['weight_mc']) != str(updated_item['weight_mc'])
            pack_changed = str(original_item['pack']) != str(updated_item['pack'])
            remark_changed = str(original_item['remark']) != str(updated_item['remark'])
            
            if quantity_changed or weight_changed or pack_changed or remark_changed:
                change_details = []
                if quantity_changed:
                    change_details.append(f"Quantity: {original_item['quantity']} → {updated_item['quantity']}")
                if weight_changed:
                    change_details.append(f"Weight MC: {original_item['weight_mc']} → {updated_item['weight_mc']}")
                if pack_changed:
                    change_details.append(f"Pack: {original_item['pack']} → {updated_item['pack']}")
                if remark_changed:
                    change_details.append(f"Remark: {original_item['remark']} → {updated_item['remark']}")
                
                changes.append({
                    'status': 'modified',
                    'fish_name': updated_item['fish_name'],
                    'packed_size': updated_item['packed_size'],
                    'old_quantity': original_item['quantity'],
                    'new_quantity': updated_item['quantity'],
                    'old_weight_mc': original_item['weight_mc'],
                    'new_weight_mc': updated_item['weight_mc'],
                    'old_pack': original_item['pack'],
                    'new_pack': updated_item['pack'],
                    'old_remark': original_item['remark'],
                    'new_remark': updated_item['remark'],
                    'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                    'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                    'changes': "; ".join(change_details)
                })
                modified_count += 1
            else:
                changes.append({
                    'status': 'unchanged',
                    'fish_name': updated_item['fish_name'],
                    'packed_size': updated_item['packed_size'],
                    'old_quantity': original_item['quantity'],
                    'new_quantity': updated_item['quantity'],
                    'old_weight_mc': original_item['weight_mc'],
                    'new_weight_mc': updated_item['weight_mc'],
                    'old_pack': original_item['pack'],
                    'new_pack': updated_item['pack'],
                    'old_remark': original_item['remark'],
                    'new_remark': updated_item['remark'],
                    'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                    'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                    'changes': 'No changes'
                })
                unchanged_count += 1
        elif updated_item and not original_item:
            # New item added
            changes.append({
                'status': 'added',
                'fish_name': updated_item['fish_name'],
                'packed_size': updated_item['packed_size'],
                'old_quantity': None,
                'new_quantity': updated_item['quantity'],
                'old_weight_mc': None,
                'new_weight_mc': updated_item['weight_mc'],
                'old_pack': None,
                'new_pack': updated_item['pack'],
                'old_remark': None,
                'new_remark': updated_item['remark'],
                'old_order_kg_per_ctn': None,
                'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                'changes': 'New item added'
            })
            added_count += 1
        elif original_item and not updated_item:
            # Item deleted
            changes.append({
                'status': 'deleted',
                'fish_name': original_item['fish_name'],
                'packed_size': original_item['packed_size'],
                'old_quantity': original_item['quantity'],
                'new_quantity': None,
                'old_weight_mc': original_item['weight_mc'],
                'new_weight_mc': None,
                'old_pack': original_item['pack'],
                'new_pack': None,
                'old_remark': original_item['remark'],
                'new_remark': None,
                'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                'new_order_kg_per_ctn': None,
                'changes': 'Item removed'
            })
            deleted_count += 1
    
    # Sort changes by status and fish name
    changes.sort(key=lambda x: (
        {'added': 0, 'modified': 1, 'deleted': 2, 'unchanged': 3}[x['status']],
        x['fish_name'].lower()
    ))
    
    return {
        'summary': {
            'added': added_count,
            'modified': modified_count,
            'deleted': deleted_count,
            'unchanged': unchanged_count
        },
        'changes': changes
    }


@app.post("/apply-order-changes")
def apply_order_changes():
    """Apply changes from comparison back to the original order file and update results."""
    try:
        original_token = request.form.get('original_token')
        batch_token = request.form.get('batch_token')
        changes_json = request.form.get('changes')
        
        if not all([original_token, batch_token, changes_json]):
            return {"success": False, "error": "Missing required parameters"}, 400
        
        # Parse changes
        try:
            changes = json.loads(changes_json)
        except:
            return {"success": False, "error": "Invalid changes data"}, 400
        
        # Get original file data from RESULT_STORE
        original_entry = RESULT_STORE.get(original_token)
        if not original_entry:
            return {"success": False, "error": "Original file not found"}, 404
        
        # Parse original file data - use stored original order data if available
        try:
            if "original_order_json" in original_entry:
                # Use stored original order data
                original_rows = eval(original_entry.get("original_order_json", b"[]"))
            else:
                # Fallback: reconstruct from processed results
                processed_rows = eval(original_entry.get("rows_json", b"[]"))
                original_rows = []
                for row in processed_rows:
                    order_row = {
                        'fish name': row.get('fish name', ''),
                        'packed size': row.get('packed size', ''),
                        'total carton': row.get('order_carton', 0),
                        'weight_mc': '',
                        'order_kg_per_ctn': row.get('order_kg_per_ctn', 0)
                    }
                    original_rows.append(order_row)
        except:
            return {"success": False, "error": "Could not parse original file data"}, 500
        
        # Apply changes to create updated order data
        updated_order_data = apply_changes_to_order_data(original_rows, changes)
        
        # Convert updated order data to proper format for compute_matches
        updated_order_rows = []
        for row in updated_order_data:
            # Convert back to original order format
            order_row = {
                'fish name': row.get('fish name', ''),
                'packed size': row.get('packed size', ''),
                'total carton': row.get('total carton', 0),
                'weight_mc': row.get('weight_mc', '')
            }
            updated_order_rows.append(order_row)
        
        # Get stock data for reprocessing
        stock_name = (original_entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        
        # Find stock file from the batch to get stock data
        session_id = get_session_id_by_batch_token(batch_token)
        if not session_id:
            return {"success": False, "error": "Batch session not found"}, 404
        
        # Get stock file data
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT stored_filename FROM uploaded_files WHERE session_id = ? AND file_type = 'stock' LIMIT 1", (session_id,))
        stock_file_row = cursor.fetchone()
        conn.close()
        
        if not stock_file_row:
            return {"success": False, "error": "Stock file not found"}, 404
        
        # Load stock data from saved file
        stock_file_path = os.path.join(UPLOAD_FOLDER, stock_file_row['stored_filename'])
        try:
            with open(stock_file_path, 'rb') as f:
                stock_ds = load_excel(f)
        except:
            return {"success": False, "error": "Could not load stock file"}, 500
        
        # Recompute matches with updated order data
        new_result_rows = compute_matches(stock_ds.rows, updated_order_rows)
        
        # Calculate new summary
        def sum_required_kg(rows):
            return round(sum(float(r.get('required_kg', 0) or 0) for r in rows), 3)

        def sum_ready_kg(rows):
            return round(sum(float(r.get('stock_carton', 0) or 0) * float(r.get('stock_kg_per_ctn', 0) or 0) for r in rows), 3)

        new_summary = {
            "total_items": len(new_result_rows),
            "full": sum(1 for r in new_result_rows if r.get('status') == 'Full'),
            "not_full": sum(1 for r in new_result_rows if r.get('status') == 'Not Full'),
            "not_have": sum(1 for r in new_result_rows if r.get('status') == 'Not have'),
            "total_kg_all": sum_required_kg(new_result_rows),
            "total_kg_full": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Full']),
            "total_kg_not_full": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Not Full']),
            "total_kg_not_have": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Not have']),
            "ready_kg": sum_ready_kg(new_result_rows)
        }
        
        # Create new Excel and PDF files
        excel_bytes = rows_to_excel_bytes(new_result_rows)
        pdf_bytes = rows_to_pdf_bytes(new_result_rows)
        
        # Update RESULT_STORE
        order_name = (original_entry.get("order_name") or b"").decode("utf-8", errors="ignore")
        order_basename = os.path.splitext(order_name)[0] if order_name else "order"
        
        RESULT_STORE[original_token] = {
            "excel": excel_bytes,
            "pdf": pdf_bytes,
            "excel_name": f"{order_basename} Calculation.xlsx",
            "pdf_name": f"{order_basename} Calculation.pdf",
            "rows_json": io.BytesIO(str(new_result_rows).encode("utf-8")).getvalue(),
            "summary_json": io.BytesIO(str(new_summary).encode("utf-8")).getvalue(),
            "stock_name": original_entry.get("stock_name", b""),
            "order_name": original_entry.get("order_name", b""),
            "original_order_json": io.BytesIO(str(updated_order_data).encode("utf-8")).getvalue(),
        }
        
        # Update database if session exists
        try:
            if session_id:
                # Find order file ID
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT id, stored_filename FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", 
                             (session_id, order_name))
                order_file_record = cursor.fetchone()
                
                if order_file_record:
                    order_file_id = order_file_record[0]
                    stored_filename = order_file_record[1]
                    
                    # Save the updated order data as a new file version
                    try:
                        # Create a new Excel file with the updated order data
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Order"
                        
                        # Add headers
                        headers = ['Fish Name', 'Packed Size', 'Pack', 'Total Carton', 'Weight MC', 'Order KG/CTN', 'Remark']
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        # Add data
                        for row_idx, order_row in enumerate(updated_order_data, 2):
                            weight_mc = order_row.get('weight_mc', '')
                            order_kg_per_ctn = order_row.get('order_kg_per_ctn', 0)
                            
                            # If weight_mc is empty but we have order_kg_per_ctn, use it
                            if not weight_mc and order_kg_per_ctn:
                                weight_mc = str(order_kg_per_ctn)
                                
                            ws.cell(row=row_idx, column=1, value=order_row.get('fish name', ''))
                            ws.cell(row=row_idx, column=2, value=order_row.get('packed size', ''))
                            ws.cell(row=row_idx, column=3, value=order_row.get('pack', ''))
                            ws.cell(row=row_idx, column=4, value=order_row.get('total carton', 0))
                            ws.cell(row=row_idx, column=5, value=weight_mc)
                            ws.cell(row=row_idx, column=6, value=order_kg_per_ctn)
                            ws.cell(row=row_idx, column=7, value=order_row.get('remark', ''))
                        
                        # Save to the same filename (overwrite)
                        updated_file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
                        wb.save(updated_file_path)
                        
                    except Exception as e:
                        print(f"Warning: Failed to save updated order file: {e}")
                    
                    # Delete old results
                    cursor.execute('DELETE FROM processing_results WHERE session_id=? AND order_file_id=?', (session_id, order_file_id))
                    # Save new results
                    save_processing_results(session_id, new_result_rows, order_file_id)
                
                # Update batch aggregated summary in processing_sessions table
                # Get all order tokens for this batch to recalculate totals
                tokens = BATCH_STORE.get(batch_token, [])
                agg_total_items = 0
                agg_full = 0
                agg_nf = 0
                agg_nh = 0
                agg_kg_all = 0
                agg_kg_full = 0
                agg_kg_nf = 0
                agg_kg_nh = 0
                
                for token in tokens:
                    entry = RESULT_STORE.get(token, {})
                    try:
                        summary = eval(entry.get("summary_json", b"{}"))
                        agg_total_items += summary.get("total_items", 0)
                        agg_full += summary.get("full", 0)
                        agg_nf += summary.get("not_full", 0)
                        agg_nh += summary.get("not_have", 0)
                        agg_kg_all += summary.get("total_kg_all", 0)
                        agg_kg_full += summary.get("total_kg_full", 0)
                        agg_kg_nf += summary.get("total_kg_not_full", 0)
                        agg_kg_nh += summary.get("total_kg_not_have", 0)
                    except:
                        pass
                
                cursor.execute(
                    '''UPDATE processing_sessions
                       SET total_items=?, full_items=?, not_full_items=?, not_have_items=?,
                           total_kg=?, full_kg=?, not_full_kg=?, not_have_kg=?
                     WHERE id=?''',
                    (
                        int(agg_total_items), int(agg_full), int(agg_nf), int(agg_nh),
                        float(agg_kg_all), float(agg_kg_full), float(agg_kg_nf), float(agg_kg_nh),
                        session_id,
                    )
                )
                
                conn.commit()
                conn.close()
        except Exception as e:
            print(f"Warning: Could not update database: {e}")
        
        # Update comparison history with applied changes
        try:
            if session_id:
                # Find order file ID
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", 
                             (session_id, order_name))
                order_file_record = cursor.fetchone()
                
                if order_file_record:
                    order_file_id = order_file_record[0]
                    # Update the most recent comparison history record with applied changes
                    cursor.execute('''UPDATE file_comparison_history 
                                     SET changes_applied = ? 
                                     WHERE session_id = ? AND order_file_id = ? AND original_token = ?
                                     ORDER BY created_at DESC LIMIT 1''',
                                 (json.dumps(changes), session_id, order_file_id, original_token))
                    conn.commit()
                
                conn.close()
        except Exception as e:
            print(f"Warning: Failed to update comparison history: {e}")
        
        return {"success": True}
        
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


def apply_changes_to_order_data(original_rows, changes):
    """Apply comparison changes to create updated order data."""
    # Start with all original data (keep unchanged items)
    updated_rows = [row.copy() for row in original_rows]
    
    # Create a lookup for original rows for quick access
    original_lookup = {}
    for i, row in enumerate(original_rows):
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        original_lookup[key] = i
    
    # Apply changes
    for change in changes:
        key = (
            canonicalize_product(change.get('fish_name', '')),
            canonicalize_pack(change.get('packed_size', ''))
        )
        
        if change['status'] == 'added':
            # Add new item to the list
            weight_mc = ''
            order_kg_per_ctn = 0
            try:
                # Try to parse kg from packed size if weight_mc is empty
                order_kg_per_ctn = parse_kg_per_carton(change['packed_size'])
            except:
                pass
            
            new_row = {
                'fish name': change['fish_name'],
                'packed size': change['packed_size'],
                'pack': change.get('new_pack', ''),
                'total carton': change['new_quantity'],
                'weight_mc': weight_mc,
                'order_kg_per_ctn': order_kg_per_ctn,
                'remark': change.get('new_remark', '')
            }
            updated_rows.append(new_row)
            
        elif change['status'] == 'modified':
            # Update existing item quantity and weight_mc
            if key in original_lookup:
                row_index = original_lookup[key]
                updated_rows[row_index]['total carton'] = change['new_quantity']
                if 'new_weight_mc' in change:
                    updated_rows[row_index]['weight_mc'] = change['new_weight_mc']
                if 'new_order_kg_per_ctn' in change:
                    updated_rows[row_index]['order_kg_per_ctn'] = change['new_order_kg_per_ctn']
                if 'new_pack' in change:
                    updated_rows[row_index]['pack'] = change['new_pack']
                if 'new_remark' in change:
                    updated_rows[row_index]['remark'] = change['new_remark']
                
        elif change['status'] == 'deleted':
            # Remove item from the list
            if key in original_lookup:
                row_index = original_lookup[key]
                # Mark for removal (we'll filter out later to maintain indices)
                updated_rows[row_index] = None
    
    # Filter out deleted items (marked as None)
    updated_rows = [row for row in updated_rows if row is not None]
    
    return updated_rows


if __name__ == "__main__":
    app.run(debug=True)


